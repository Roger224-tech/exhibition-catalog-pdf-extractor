"""
Step 5: 端到端管线 + Excel导出
==============================
串联 Step1-4，从 PDF 一次生成最终 Excel 文件。

功能:
  1. 整合 PDF解析 → 字段抽取 → 图片关联 全流程
  2. 导出 Excel (.xlsx)，支持图片嵌入或路径引用
  3. 生成处理报告 (JSON + 文本摘要)
  4. 性能计时，对比验收标准

用法:
    python step5_e2e_pipeline.py ../test_pdfs/sample_a.pdf
    python step5_e2e_pipeline.py ../test_pdfs/sample_a.pdf --image-mode path
    python step5_e2e_pipeline.py ../test_pdfs/  # 批量处理
"""

import fitz
import json
import os
import sys
import time
import argparse
from pathlib import Path
from collections import defaultdict

try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

# 导入各模块
sys.path.insert(0, os.path.dirname(__file__))

from step1_pdf_parse import parse_pdf, export_images
from step3_field_extraction import FieldExtractor
from step4_image_matching import match_images_to_products


# ═════════════════════════════════════════════════════
#  Excel 导出
# ═════════════════════════════════════════════════════

def export_to_excel(
    products: list,
    image_matches: dict,
    images_dir: str,
    output_path: str,
    image_mode: str = "embed",
    thumbnail_size: tuple = (120, 120),
):
    """
    生成最终 Excel 文件

    参数:
      products:     字段抽取结果列表
      image_matches: 图片关联结果 {page: [matches]}
      images_dir:   图片文件目录
      output_path:  Excel输出路径
      image_mode:   "embed" 嵌入缩略图 | "path" 引用路径
      thumbnail_size: 缩略图最大尺寸
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    try:
        from PIL import Image as PILImage
        import io
        HAS_PIL = True
    except ImportError:
        HAS_PIL = False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "产品目录"

    # ── 表头 ──
    HEADERS = [
        ("页码", 6),
        ("品牌 / 制造商", 14),
        ("车型适配", 22),
        ("产品编号 / OE号", 18),
        ("描述一", 28),
        ("描述二", 30),
        ("描述三 / 规格", 24),
        ("价格", 12),
        ("原厂参考号", 16),
        ("每包数量", 10),
        ("产品图片", 15 if image_mode == "path" else 18),
        ("AI置信度", 10),
        ("备注", 15),
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, (header, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # 冻结表头
    ws.freeze_panes = "A2"

    # ── 数据行 ──
    low_conf_fill = PatternFill("solid", fgColor="FFF2CC")  # 低置信度黄色

    for row_idx, product in enumerate(products, 2):
        ws.cell(row=row_idx, column=1, value=product.get("page", "")).border = thin_border

        # 品牌
        ws.cell(row=row_idx, column=2,
                value=_field_val(product, "brand")).border = thin_border
        _mark_low_conf(ws, row_idx, 2, product.get("brand", {}), low_conf_fill)

        # 车型适配
        ws.cell(row=row_idx, column=3,
                value=_field_val(product, "vehicle_fitment")).border = thin_border

        # OE号
        ws.cell(row=row_idx, column=4,
                value=_field_val(product, "oe_number")).border = thin_border
        _mark_low_conf(ws, row_idx, 4, product.get("oe_number", {}), low_conf_fill)

        # 描述一
        ws.cell(row=row_idx, column=5,
                value=_field_val(product, "description_1")).border = thin_border

        # 描述二
        ws.cell(row=row_idx, column=6,
                value=_field_val(product, "description_2")).border = thin_border

        # 描述三/规格
        ws.cell(row=row_idx, column=7,
                value=_field_val(product, "description_3")).border = thin_border

        # 价格
        ws.cell(row=row_idx, column=8,
                value=_field_val(product, "price")).border = thin_border

        # 原厂参考号
        ws.cell(row=row_idx, column=9,
                value=_field_val(product, "oem_ref")).border = thin_border

        # 每包数量
        ws.cell(row=row_idx, column=10,
                value=_field_val(product, "pack_qty")).border = thin_border

        # 图片列
        img_col = 11
        ws.cell(row=row_idx, column=img_col).border = thin_border

        # 置信度
        conf_avg = product.get("confidence_avg", 0)
        conf_cell = ws.cell(row=row_idx, column=12, value=f"{conf_avg:.0%}")
        conf_cell.border = thin_border
        conf_cell.alignment = Alignment(horizontal="center")
        if conf_avg < 0.70:
            conf_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # 红色警示

        # 备注
        ws.cell(row=row_idx, column=13).border = thin_border

        # ── 图片嵌入 ──
        if image_mode == "embed" and HAS_PIL:
            # 查找该产品关联的图片
            page_num = product.get("page", 0)
            page_matches = image_matches.get(str(page_num), [])
            prod_idx = row_idx - 2  # 产品在列表中的索引
            for match in page_matches:
                if match.get("product_idx") == prod_idx:
                    img_idx = match.get("image_idx", 0)
                    # 构建图片文件名
                    img_filename = f"p{page_num:03d}_xref{img_idx}"
                    # 在images_dir中查找
                    found_img = _find_image(images_dir, img_filename)
                    if found_img and Path(found_img).exists():
                        try:
                            thumb = _create_thumbnail(found_img, thumbnail_size)
                            if thumb:
                                xl_img = XLImage(thumb)
                                # 缩放到合适大小
                                xl_img.width, xl_img.height = min(thumb.size[0], thumbnail_size[0]), min(thumb.size[1], thumbnail_size[1])
                                ws.add_image(xl_img, f"{get_column_letter(img_col)}{row_idx}")
                                ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 15, thumbnail_size[1] * 0.8)
                        except Exception:
                            pass
                    break

            # 如果没找到嵌入图片，用路径模式
            if image_mode == "path":
                img_path = _find_image(images_dir, f"p{page_num:03d}")
                if img_path:
                    ws.cell(row=row_idx, column=img_col, value=img_path)

        elif image_mode == "path":
            page_num = product.get("page", 0)
            img_path = _find_image(images_dir, f"p{page_num:03d}")
            if img_path:
                ws.cell(row=row_idx, column=img_col, value=img_path)

    # 自适应行高
    for row in range(2, len(products) + 2):
        if ws.row_dimensions[row].height is None:
            ws.row_dimensions[row].height = 20

    # 保存
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def _field_val(product: dict, field: str) -> str:
    """从产品字典中安全取字段值"""
    val = product.get(field, {})
    if isinstance(val, dict):
        return val.get("value", "")
    return str(val) if val else ""


def _mark_low_conf(ws, row: int, col: int, field_data: dict, fill):
    """低置信度字段标黄"""
    conf = field_data.get("confidence", 1.0)
    if conf < 0.70 and conf > 0:
        ws.cell(row=row, column=col).fill = fill


def _find_image(images_dir: str, prefix: str) -> str | None:
    """在图片目录中查找匹配前缀的图片"""
    if not images_dir or not Path(images_dir).exists():
        return None
    for f in Path(images_dir).iterdir():
        if f.name.startswith(prefix):
            return str(f)
    return None


def _create_thumbnail(image_path: str, max_size: tuple) -> object | None:
    """创建缩略图 (返回 BytesIO)"""
    try:
        from PIL import Image as PILImage
        import io
        img = PILImage.open(image_path)
        img.thumbnail(max_size)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        # 返回PIL Image对象供openpyxl使用
        return PILImage.open(buf)
    except Exception:
        return None


# ═════════════════════════════════════════════════════
#  处理报告
# ═════════════════════════════════════════════════════

def generate_report(
    pdf_name: str,
    parsed_data: dict,
    all_products: list,
    image_matches: dict,
    excel_path: str,
    total_time: float,
    output_dir: str,
) -> dict:
    """生成处理报告"""
    total_images = sum(
        len(p.get("images", [])) for p in parsed_data.get("pages", [])
    )
    total_matches = sum(len(v) for v in image_matches.values())

    scanned_pages = [
        p["page_num"] for p in parsed_data.get("pages", [])
        if p.get("is_scanned") and not p.get("ocr_applied")
    ]

    low_conf_products = [
        p for p in all_products
        if p.get("confidence_avg", 0) < 0.5
    ]

    avg_confidence = (
        sum(p.get("confidence_avg", 0) for p in all_products) / max(len(all_products), 1)
    )

    report = {
        "file": pdf_name,
        "processing_time_seconds": round(total_time, 2),
        "total_pages": parsed_data.get("total_pages", 0),
        "scanned_pages": scanned_pages,
        "total_products": len(all_products),
        "total_images": total_images,
        "matched_images": total_matches,
        "avg_confidence": round(avg_confidence, 2),
        "low_confidence_products": len(low_conf_products),
        "output_excel": excel_path,
        "warnings": [],
    }

    # 添加警告
    if scanned_pages:
        report["warnings"].append(f"{len(scanned_pages)}页扫描件未处理")
    if avg_confidence < 0.6:
        report["warnings"].append(f"平均置信度偏低 ({avg_confidence:.0%})")
    if total_matches < total_images:
        report["warnings"].append(f"{total_images - total_matches}张图片未关联到产品")

    # 保存报告
    report_path = os.path.join(output_dir, f"{pdf_name}_report.json")
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    # 文本摘要
    summary_path = os.path.join(output_dir, f"{pdf_name}_report.txt")
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write(f"处理报告: {pdf_name}\n")
        f.write(f"{'='*40}\n")
        f.write(f"总页数:            {report['total_pages']}\n")
        f.write(f"扫描件页数:        {len(scanned_pages)}\n")
        f.write(f"提取产品数:        {report['total_products']}\n")
        f.write(f"总图片数:          {report['total_images']}\n")
        f.write(f"已关联图片:        {report['matched_images']}\n")
        f.write(f"平均置信度:        {avg_confidence:.0%}\n")
        f.write(f"低置信度产品:      {len(low_conf_products)}\n")
        f.write(f"处理耗时:          {total_time:.1f} 秒\n")
        f.write(f"输出Excel:         {excel_path}\n")
        if report["warnings"]:
            f.write(f"\n⚠ 告警:\n")
            for w in report["warnings"]:
                f.write(f"  - {w}\n")

    return report


# ═════════════════════════════════════════════════════
#  主入口
# ═════════════════════════════════════════════════════

def process_single_pdf(pdf_path: str, output_dir: str, image_mode: str = "embed", ocr_engine=None) -> dict:
    """处理单个PDF文件"""
    print(f"\n{'='*60}")
    print(f"  处理: {Path(pdf_path).name}")
    print(f"{'='*60}")

    pdf_name = Path(pdf_path).stem
    step_output = os.path.join(output_dir, pdf_name)
    os.makedirs(step_output, exist_ok=True)

    t_start = time.time()

    # ── Step 1: PDF解析 ──
    print(f"  [1/4] PDF解析...")
    parsed = parse_pdf(pdf_path, ocr_engine=ocr_engine)
    doc = fitz.open(pdf_path)
    img_dir, exported_imgs = export_images(doc, parsed["pages"], step_output, pdf_name)
    doc.close()

    # ── Step 2+3: 字段抽取 ──
    print(f"  [2/4] 字段抽取...")
    extractor = FieldExtractor()
    all_products = []

    for page_data in parsed["pages"]:
        # 跳过扫描件（未OCR处理的），但保留OCR处理过的页面
        if page_data.get("is_scanned") and not page_data.get("ocr_applied"):
            continue
        products = extractor.extract_from_page({
            "text_blocks": page_data.get("text_blocks", []),
            "images": page_data.get("images", []),
            "card_regions": [],
            "page_size": page_data.get("page_size", [0, 0, 600, 800]),
        }, page_data["page_num"])
        all_products.extend(products)

    # ── Step 4: 图片关联 ──
    print(f"  [3/4] 图片关联...")
    images_by_page = defaultdict(list)
    for page_data in parsed["pages"]:
        images_by_page[page_data["page_num"]] = page_data.get("images", [])

    all_matches = {}
    for page_num, imgs in images_by_page.items():
        page_size = (600, 800)
        matches = match_images_to_products(all_products, imgs, page_size, page_num)
        if matches:
            all_matches[str(page_num)] = matches

    # ── 保存中间结果 ──
    fields_path = os.path.join(step_output, f"{pdf_name}_fields.json")
    with open(fields_path, "w", encoding="utf-8") as f:
        json.dump({"source_file": pdf_name, "total_products": len(all_products),
                    "products": all_products}, f, ensure_ascii=False, indent=2, default=str)

    # ── 导出 Excel ──
    print(f"  [4/4] Excel导出...")
    excel_path = os.path.join(step_output, f"{pdf_name}_output.xlsx")
    export_to_excel(all_products, all_matches, img_dir, excel_path, image_mode=image_mode)

    total_time = time.time() - t_start

    # ── 报告 ──
    report = generate_report(
        pdf_name, parsed, all_products, all_matches,
        excel_path, total_time, step_output,
    )

    # ── 控制台输出 ──
    print(f"\n  ── 处理结果 ──")
    print(f"  总页数:      {report['total_pages']}")
    print(f"  产品数:      {report['total_products']}")
    print(f"  图片数:      {report['total_images']} (已关联: {report['matched_images']})")
    print(f"  平均置信度:  {report['avg_confidence']:.0%}")
    print(f"  耗时:        {total_time:.1f} 秒")
    print(f"  输出:        {excel_path}")
    if report["warnings"]:
        for w in report["warnings"]:
            print(f"  [WARN] {w}")
    print()

    return report


def main():
    parser = argparse.ArgumentParser(description="Step 5: 端到端管线 + Excel导出")
    parser.add_argument("input_path", help="PDF文件或文件夹路径")
    parser.add_argument("-o", "--output", default="step5_output", help="输出目录")
    parser.add_argument("--image-mode", choices=["embed", "path"], default="embed",
                        help="图片处理模式: embed(缩略图嵌入) | path(路径引用)")
    parser.add_argument("--ocr", choices=["paddleocr", "tesseract"], default="paddleocr",
                        help="OCR引擎 (默认: paddleocr)")
    parser.add_argument("--no-ocr", action="store_true", help="禁用OCR (仅处理文字型PDF)")
    args = parser.parse_args()

    input_path = Path(args.input_path)
    pdf_files = []

    if input_path.is_dir():
        pdf_files = sorted(input_path.glob("*.pdf"))
        if not pdf_files:
            print(f"错误: 目录 '{input_path}' 中未找到PDF文件")
            sys.exit(1)
    elif input_path.suffix.lower() == ".pdf":
        pdf_files = [input_path]
    else:
        print(f"错误: '{input_path}' 不是PDF文件")
        sys.exit(1)

    os.makedirs(args.output, exist_ok=True)
    all_reports = []

    # 初始化OCR引擎
    ocr_engine = None
    if not args.no_ocr:
        from ocr_engine import OCREngine
        print(f"初始化OCR引擎: {args.ocr}")
        ocr_engine = OCREngine(engine=args.ocr)
        if not ocr_engine.is_available:
            print("[WARN] OCR引擎不可用，仅处理文字型PDF页面")
            ocr_engine = None

    for pdf_path in pdf_files:
        if not pdf_path.exists():
            print(f"跳过不存在的文件: {pdf_path}")
            continue
        report = process_single_pdf(str(pdf_path), args.output, args.image_mode, ocr_engine=ocr_engine)
        all_reports.append(report)

    # 批量汇总
    if len(all_reports) > 1:
        print(f"\n{'='*60}")
        print(f"  批量处理汇总 ({len(all_reports)} 个文件)")
        print(f"{'='*60}")
        total_prods = sum(r["total_products"] for r in all_reports)
        total_imgs = sum(r["total_images"] for r in all_reports)
        total_time = sum(r["processing_time_seconds"] for r in all_reports)
        avg_conf = sum(r["avg_confidence"] for r in all_reports) / len(all_reports)
        print(f"  总产品数:      {total_prods}")
        print(f"  总图片数:      {total_imgs}")
        print(f"  总耗时:        {total_time:.1f} 秒")
        print(f"  平均置信度:    {avg_conf:.0%}")
        print()

        # 保存汇总报告
        summary = {
            "total_files": len(all_reports),
            "total_products": total_prods,
            "total_images": total_imgs,
            "total_time_seconds": round(total_time, 2),
            "avg_confidence": round(avg_conf, 2),
            "per_file": all_reports,
        }
        summary_path = os.path.join(args.output, "_batch_summary.json")
        with open(summary_path, "w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2, default=str)
        print(f"汇总报告: {summary_path}")


if __name__ == "__main__":
    main()
